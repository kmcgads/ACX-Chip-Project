"""
csvvolcont.py
─────────────────────────────────────────────────────────────────────────────
The original code for this chip was written in C++ by ACX Instruments and
later adapted for Python using ctypes. To use this chip, the user must
purchase the hardware from ACX Instruments. ACX provides the required starter
software and DLL files with the purchased device. Because the DLL is
proprietary company software, I cannot share the actual DLL file or its file
path. The placeholder below represents where the ACX-provided DLL would be
loaded.

─── Usage from master script ──────────────────────────────────────────────────

    import csvvolcont

    csvvolcont.initialize()   # call ONCE at the start — opens USB, sets power/voltage
    ...
    csvvolcont.main()         # call each trial — loads CSV widths, runs mix sequence
    ...

IMPORTANT — Connection management
──────────────────────────────────
initialize() opens the USB connection. It must NOT be called again between
trials — the connection stays open for the entire experiment. main() does NOT
open or close the USB connection.
────────────────────────────────────────────────────────────────────────────────
"""

import ctypes
from ctypes import Structure
import time
import os
import openpyxl

# ── DLL load ──────────────────────────────────────────────────────────────────

os.add_dll_directory(r"C:\Users\klmcg\Downloads\ACX_pythonSDK v1.2 3\ACX_pythonSDK\windows")
microfluidics = ctypes.CDLL(
    r"C:\Users\klmcg\Downloads\ACX_pythonSDK v1.2 3\ACX_pythonSDK\windows\DLLTest.dll"
)


class Drop(Structure):
    _fields_ = [
        ("height", ctypes.c_int),
        ("width",  ctypes.c_int),
        ("row",    ctypes.c_int),
        ("col",    ctypes.c_int),
    ]


def activate(drops, debug_label=""):
    """Send an electrode activation command over the existing open connection."""
    n   = len(drops)
    arr = (Drop * n)(*drops)
    print(f"\n--- ACTIVATE CALL: {debug_label} ---")
    print(f"    Total drops sent to device: {n}")
    for idx, d in enumerate(drops):
        print(
            f"    Drop[{idx}]: "
            f"row={d.row}, col={d.col}, "
            f"height={d.height}, width={d.width} "
            f"| covers rows {d.row}–{d.row + d.height - 1}, "
            f"cols {d.col}–{d.col + d.width - 1}"
        )
    microfluidics.ActivateElec(128, 128, n, arr)
    time.sleep(0.5)


# ── CSV loader ────────────────────────────────────────────────────────────────

CSV_PATH = r"C:\Users\klmcg\OneDrive\Documents\colormixcsv.xlsx"

def load_piece_widths(filepath=CSV_PATH):
    """
    Read piece end-widths from row 2 of the Excel file.
    Expected columns: piece_1 width | piece_2 width | piece_3 width
    Height is always MAIN_H=10 and is NOT read from the file.
    Returns (piece1_end_w, piece2_end_w, piece3_end_w) as ints.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    piece1_end_w = int(ws.cell(row=2, column=1).value)
    piece2_end_w = int(ws.cell(row=2, column=2).value)
    piece3_end_w = int(ws.cell(row=2, column=3).value)
    print(f"\n[CSV] Loaded piece end-widths from: {filepath}")
    print(f"      piece_1 width={piece1_end_w}, piece_2 width={piece2_end_w}, piece_3 width={piece3_end_w}")
    print(f"      Height is fixed at {MAIN_H} for all drops.\n")
    return piece1_end_w, piece2_end_w, piece3_end_w


# ── Constants ─────────────────────────────────────────────────────────────────

MAIN_H          = 10
MAIN_W          = 15
MAIN_COL        = 2
DROP1_ROW       = 75
DROP2_ROW       = 115
DROP3_ROW       = 25

PIECE_START_COL = 30
PIECE_START_W   = 10
STRETCH_STEPS   = 25
PIECE_FINAL_COL = PIECE_START_COL + STRETCH_STEPS  # col=55

MEETING_ROW     = DROP1_ROW        # row=55
MEETING_COL     = PIECE_FINAL_COL  # col=55


# ── One-time USB initialization — call from master script at startup ───────────

def initialize():
    """
    Opens the USB connection, sets power and voltage.
    Call this ONCE at the start of the experiment — NOT between trials.
    The connection stays open until the experiment ends.
    """
    print("\n[csvvolcont] Initializing USB connection...")
    microfluidics.InitUSB()
    res = microfluidics.OpenUSB()
    if res:
        print("[csvvolcont] USB opened successfully.")
    else:
        raise RuntimeError(
            "[csvvolcont] OpenUSB() failed — device may already be open, "
            "not connected, or in use by another process."
        )

    microfluidics.SetPower(True)
    print("[csvvolcont] Power on.")

    # Explicitly cast to c_int so ctypes passes the correct C type to the DLL
    microfluidics.SetVolt(
        ctypes.c_int(45), ctypes.c_int(45), ctypes.c_int(45),
        ctypes.c_int(0),  ctypes.c_int(0),  ctypes.c_int(0),
        ctypes.c_int(0),  ctypes.c_int(0),  ctypes.c_int(0),
    )
    print("[csvvolcont] Voltage set.")

    time.sleep(0.3)   # small delay so DLL has time to apply before querying

    v1 = ctypes.c_int(0); v2 = ctypes.c_int(0); v3 = ctypes.c_int(0)
    v4 = ctypes.c_int(0); v5 = ctypes.c_int(0); v6 = ctypes.c_int(0)
    v7 = ctypes.c_int(0); v8 = ctypes.c_int(0); v9 = ctypes.c_int(0)
    microfluidics.InquireVolt(
        ctypes.byref(v1), ctypes.byref(v2), ctypes.byref(v3),
        ctypes.byref(v4), ctypes.byref(v5), ctypes.byref(v6),
        ctypes.byref(v7), ctypes.byref(v8), ctypes.byref(v9)
    )
    print(f"[csvvolcont] Voltages confirmed: "
          f"{v1.value} {v2.value} {v3.value} {v4.value} {v5.value} "
          f"{v6.value} {v7.value} {v8.value} {v9.value}")
    # Allow ±2V tolerance — the DLL may round internally (e.g. 45 → 46)
    VOLT_TARGET    = 45
    VOLT_TOLERANCE = 2
    for label, v in [("CH1", v1), ("CH2", v2), ("CH3", v3)]:
        if abs(v.value - VOLT_TARGET) > VOLT_TOLERANCE:
            raise RuntimeError(
                f"[csvvolcont] Voltage verification failed on {label} — "
                f"expected ~{VOLT_TARGET}V, got {v.value}V. Check DLL argument types."
            )
    print(f"[csvvolcont] Voltage verification passed (tolerance ±{VOLT_TOLERANCE}V).")
    print("[csvvolcont] Initialization complete.\n")


# ── Helpers ───────────────────────────────────────────────────────────────────

def held_drops(held_items):
    """
    held_items: list of (row, main_col, piece_end_w) tuples for each drop currently held.
    Returns a list of Drop objects — one main body and one piece per held drop.
    """
    drops = []
    for r, mc, piece_end_w in held_items:
        drops.append(Drop(MAIN_H, MAIN_W,       r, mc))
        drops.append(Drop(MAIN_H, piece_end_w,  r, PIECE_FINAL_COL))
    return drops


def split_and_move(row, label, held_items, piece_end_w, start_col=None):
    """
    Loads, stretches, splits, and moves a drop's piece to PIECE_FINAL_COL.
    Each sub-step requires the user to press Enter before proceeding.
    """
    if start_col is None:
        start_col = MAIN_COL

    neck_start = start_col + MAIN_W
    neck_gap   = PIECE_START_COL - neck_start

    # Step 1: Load
    input(f"\n>>> [{label}] Ready to LOAD drop at row={row}, col={start_col} — press Enter")
    print(f"[{label}] Loading drop...")
    activate(
        held_drops(held_items) + [Drop(MAIN_H, 20, row, start_col)],
        debug_label=f"{label} LOAD"
    )
    time.sleep(1)

    # Step 2: Stretch
    input(f"\n>>> [{label}] Ready to STRETCH width 20 → 35 — press Enter")
    print(f"[{label}] Stretching...")
    time.sleep(2)
    for i in range(1, 16):
        activate(
            held_drops(held_items) + [Drop(MAIN_H, 20 + i, row, start_col)],
            debug_label=f"{label} STRETCH width={20 + i}"
        )

    # Step 3: Open split gap
    input(f"\n>>> [{label}] Ready to SPLIT (open gap {neck_gap} steps) — press Enter")
    print(f"[{label}] Opening split gap...")
    time.sleep(2)
    for gap in range(neck_gap + 1):
        bridge_col   = neck_start + gap
        bridge_width = neck_gap   - gap
        if bridge_width > 0:
            activate(
                held_drops(held_items) + [
                    Drop(MAIN_H, MAIN_W,        row, start_col),
                    Drop(MAIN_H, bridge_width,  row, bridge_col),
                    Drop(MAIN_H, PIECE_START_W, row, PIECE_START_COL),
                ],
                debug_label=f"{label} SPLIT gap={gap}"
            )
        else:
            activate(
                held_drops(held_items) + [
                    Drop(MAIN_H, MAIN_W,        row, start_col),
                    Drop(MAIN_H, PIECE_START_W, row, PIECE_START_COL),
                ],
                debug_label=f"{label} SPLIT gap fully open"
            )

    # Step 4: Move piece
    input(f"\n>>> [{label}] Ready to MOVE piece right, pinching {PIECE_START_W} → {piece_end_w} — press Enter")
    print(f"[{label}] Moving piece...")
    time.sleep(2)
    for i in range(1, STRETCH_STEPS + 1):
        current_col   = PIECE_START_COL + i
        current_width = round(PIECE_START_W - (PIECE_START_W - piece_end_w) * i / STRETCH_STEPS)
        activate(
            held_drops(held_items) + [
                Drop(MAIN_H, MAIN_W,         row, start_col),
                Drop(MAIN_H, current_width,  row, current_col),
            ],
            debug_label=f"{label} MOVE col={current_col} width={current_width}"
        )

    # Step 5: Deactivate neck
    input(f"\n>>> [{label}] Ready to DEACTIVATE neck — press Enter")
    print(f"[{label}] Deactivating neck...")
    neck_end = PIECE_FINAL_COL - 1
    time.sleep(2)
    for release_col in range(neck_end, neck_start - 1, -1):
        bridge_width = release_col - neck_start
        if bridge_width > 0:
            activate(
                held_drops(held_items) + [
                    Drop(MAIN_H, MAIN_W,        row, start_col),
                    Drop(MAIN_H, bridge_width,  row, neck_start),
                    Drop(MAIN_H, piece_end_w,   row, PIECE_FINAL_COL),
                ],
                debug_label=f"{label} DEACTIVATE col={release_col}"
            )
        else:
            activate(
                held_drops(held_items) + [
                    Drop(MAIN_H, MAIN_W,       row, start_col),
                    Drop(MAIN_H, piece_end_w,  row, PIECE_FINAL_COL),
                ],
                debug_label=f"{label} DEACTIVATE FINAL"
            )

    print(f"[{label}] Split complete.")
    time.sleep(1)


def move_pieces_to_meet(piece1_end_w, piece2_end_w, piece3_end_w):
    """
    Moves all three pieces simultaneously toward MEETING_ROW, MEETING_COL,
    merges them, then runs the full mixing sequence.
    """
    steps_to_meet = DROP2_ROW - MEETING_ROW   # 50 steps

    input(f"\n>>> All three drops split — ready to MOVE PIECES TO MEET at row={MEETING_ROW}, col={MEETING_COL} — press Enter")
    print(f"[Mix] Moving pieces to meet...")
    time.sleep(1)
    for i in range(1, steps_to_meet + 1):
        piece2_row = DROP2_ROW - i
        piece3_row = DROP3_ROW + i
        activate(
            [
                Drop(MAIN_H, MAIN_W,       DROP1_ROW,   MAIN_COL),
                Drop(MAIN_H, MAIN_W,       DROP2_ROW,   MAIN_COL),
                Drop(MAIN_H, MAIN_W,       DROP3_ROW,   MAIN_COL),
                Drop(MAIN_H, piece1_end_w, MEETING_ROW, MEETING_COL),
                Drop(MAIN_H, piece2_end_w, piece2_row,  MEETING_COL),
                Drop(MAIN_H, piece3_end_w, piece3_row,  MEETING_COL),
            ],
            debug_label=f"MOVE TO MEET step={i}"
        )

    # Merge
    input(f"\n>>> Pieces at meeting point — ready to MERGE — press Enter")
    print("[Mix] Merging pieces...")
    time.sleep(1)
    activate(
        [
            Drop(MAIN_H, MAIN_W,       DROP1_ROW,   MAIN_COL),
            Drop(MAIN_H, MAIN_W,       DROP2_ROW,   MAIN_COL),
            Drop(MAIN_H, MAIN_W,       DROP3_ROW,   MAIN_COL),
            Drop(MAIN_H, piece1_end_w, MEETING_ROW, MEETING_COL),
        ],
        debug_label="MERGE all three pieces"
    )

    # Mixing sequence
    all_mains = [
        Drop(MAIN_H, MAIN_W, DROP1_ROW, MAIN_COL),
        Drop(MAIN_H, MAIN_W, DROP2_ROW, MAIN_COL),
        Drop(MAIN_H, MAIN_W, DROP3_ROW, MAIN_COL),
    ]
    H, W = MAIN_H, MAIN_W
    r, c = MEETING_ROW, MEETING_COL

    input(f"\n>>> Pieces merged — ready to begin MIXING SEQUENCE — press Enter")
    print("\n[Mix] Running mixing sequence...")

    print("  Mix pass 1: right 30...")
    for i in range(1, 31):      activate(all_mains + [Drop(H, W, r, c + i)],       debug_label=f"MIX pass1 right i={i}")
    for i in range(30, -1, -1): activate(all_mains + [Drop(H, W, r, c + i)],       debug_label=f"MIX pass1 back i={i}")

    print("  Mix pass 2: diagonal up-right 20...")
    for i in range(1, 21):      activate(all_mains + [Drop(H, W, r - i, c + i)],   debug_label=f"MIX pass2 out i={i}")
    for i in range(20, -1, -1): activate(all_mains + [Drop(H, W, r - i, c + i)],   debug_label=f"MIX pass2 back i={i}")

    print("  Mix pass 3: right-then-down...")
    for i in range(1, 31):      activate(all_mains + [Drop(H, W, r, c + i)],       debug_label=f"MIX pass3 right i={i}")
    for i in range(1, 21):      activate(all_mains + [Drop(H, W, r + i, c + 30)],  debug_label=f"MIX pass3 down i={i}")
    for i in range(20, -1, -1): activate(all_mains + [Drop(H, W, r + i, c + i)],   debug_label=f"MIX pass3 back i={i}")

    print("  Mix pass 4: right 30 double-pass...")
    for _ in range(2):
        for i in range(1, 31):      activate(all_mains + [Drop(H, W, r, c + i)],   debug_label=f"MIX pass4 right i={i}")
        for i in range(30, -1, -1): activate(all_mains + [Drop(H, W, r, c + i)],   debug_label=f"MIX pass4 back i={i}")

    print("  Mix pass 5: diagonal down-right 15...")
    for i in range(1, 16):      activate(all_mains + [Drop(H, W, r + i, c + i)],   debug_label=f"MIX pass5 out i={i}")
    for i in range(15, -1, -1): activate(all_mains + [Drop(H, W, r + i, c + i)],   debug_label=f"MIX pass5 back i={i}")

    print("  Mix pass 6: clockwise rectangle 30x20...")
    for i in range(1, 31):      activate(all_mains + [Drop(H, W, r,      c + i)],  debug_label=f"MIX pass6 top i={i}")
    for i in range(1, 21):      activate(all_mains + [Drop(H, W, r + i,  c + 30)], debug_label=f"MIX pass6 right i={i}")
    for i in range(30, -1, -1): activate(all_mains + [Drop(H, W, r + 20, c + i)],  debug_label=f"MIX pass6 bottom i={i}")
    for i in range(20, -1, -1): activate(all_mains + [Drop(H, W, r + i,  c)],      debug_label=f"MIX pass6 left i={i}")

    print("  [Mix] Mixing complete.")


# ── Per-trial entry point — called by master script each trial ─────────────────

def main():
    """
    Runs one full mix trial: loads CSV widths, splits all three drops,
    moves pieces to meet, merges, and mixes.

    Does NOT open or close the USB connection — initialize() handles that.
    Safe to call multiple times in a loop.
    """
    piece1_end_w, piece2_end_w, piece3_end_w = load_piece_widths()

    # Drop 1
    split_and_move(
        row=DROP1_ROW,
        label="Drop 1 (row=65)",
        held_items=[],
        piece_end_w=piece1_end_w,
        start_col=MAIN_COL
    )

    # Drop 2
    split_and_move(
        row=DROP2_ROW,
        label="Drop 2 (row=115)",
        held_items=[(DROP1_ROW, MAIN_COL, piece1_end_w)],
        piece_end_w=piece2_end_w,
        start_col=MAIN_COL
    )

    # Drop 3
    split_and_move(
        row=DROP3_ROW,
        label="Drop 3 (row=10)",
        held_items=[(DROP1_ROW, MAIN_COL, piece1_end_w), (DROP2_ROW, MAIN_COL, piece2_end_w)],
        piece_end_w=piece3_end_w,
        start_col=MAIN_COL
    )

    # Move, merge, mix
    move_pieces_to_meet(piece1_end_w, piece2_end_w, piece3_end_w)


# ── Standalone entry point (manual testing only) ──────────────────────────────

if __name__ == "__main__":
    microfluidics.InitUSB()
    res = microfluidics.OpenUSB()
    if res:
        input("Open successfully — press Enter to continue")
    else:
        input("Open failed — press Enter to exit")
        raise SystemExit(1)

    microfluidics.SetPower(True)
    input("Power on completed — press Enter to set voltage")

    microfluidics.SetVolt(45, 45, 45, 0, 0, 0, 0, 0, 0)
    input("Voltage set — press Enter to run")

    main()

    input(">>> Sequence complete -- press Enter to shut down")