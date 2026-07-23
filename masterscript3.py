"""
run_experiment.py
─────────────────
Bayesian-optimized closed-loop color mixing experiment on DMF chip.

Each trial:
  Step 1 — Bayesian suggests reservoir widths → writes to colormixcsv.xlsx
  Step 2 — csvvolcont mixes the drop on chip
  Step 3 — Camera measures the resulting color
  Step 4 — Bayesian scores it (DeltaE vs target), updates its model
  Step 5 — Drop moved to graveyard, reservoirs reloaded
  Repeat up to N_CALLS times, stops early if DeltaE < 2.0.

Place in: C:\\Users\\klmcg\\SULIProj\\ACX-CHIP-PROJECT\\
Run with: python run_experiment.py
"""

import sys
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

import csvvolcont
import cleanreload
from camera import CameraInterface
from bayesopttest1 import BlindOptimizer, random_vivid_target_color

# ── Config ────────────────────────────────────────────────────────────────────

CAMERA_INDEX  = 1      # Change if microscope is on a different index
RANDOM_SEED   = None   # None = different random target color every run
N_CALLS       = 5      # Number of trials to run (change this to control trials)

# Must match the path csvvolcont.load_piece_widths() reads from
COLOR_MIX_XLSX = r"C:\Users\klmcg\OneDrive\Documents\colormixcsv.xlsx"


# ── Camera helper ─────────────────────────────────────────────────────────────

def capture_color(cam: CameraInterface) -> str:
    """Take a picture and return the measured hex color string."""
    print("\n  [Camera] Taking picture...")
    image_path, frame = cam.take_picture()
    print(f"  [Camera] Image saved: {image_path}")
    color_result = cam.detect_drop_color(frame)
    print(f"  [Camera] Measured RGB : {color_result['rgb']}")
    print(f"  [Camera] Measured hex : {color_result['hex']}")
    return color_result['hex']


# ── Width writer — writes optimizer widths directly to the xlsx csvvolcont reads ──

def write_widths_to_xlsx(w1: int, w2: int, w3: int, path: str = COLOR_MIX_XLSX) -> None:
    """
    Writes the Bayesian-suggested widths into colormixcsv.xlsx so that
    csvvolcont.load_piece_widths() picks up the correct values each trial.
    Overwrites row 2, columns 1-3 in place.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.cell(row=2, column=1).value = w1
    ws.cell(row=2, column=2).value = w2
    ws.cell(row=2, column=3).value = w3
    wb.save(path)
    print(f"  [XLSX] Widths written → piece_1={w1}, piece_2={w2}, piece_3={w3}  ({path})")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 65)
    print("  BAYESIAN OPTIMIZATION EXPERIMENT — DMF Chip")
    print("=" * 65)

    # ── Setup ──────────────────────────────────────────────────────────────
    target = random_vivid_target_color(seed=RANDOM_SEED)

    print(f"\n  Target color  : {target.hex}")
    print(f"  Target RGB    : ({target.r}, {target.g}, {target.b})")
    print(f"  Scoring metric: CIEDE2000 DeltaE  (< 2.0 = visually identical)")

    cam = CameraInterface(camera_address=CAMERA_INDEX)
    opt = BlindOptimizer(target, n_calls=N_CALLS)

    print(f"  Trials        : {opt.n_calls}  |  Trial 1 = PRESET (5,5,5), rest = GP-guided")

    # ── One-time USB initialization — connection stays open for all trials ──
    print(f"\n  [Setup] Initializing chip connection (once)...")
    csvvolcont.initialize()

    # ── Trial loop ─────────────────────────────────────────────────────────
    for trial in range(opt.n_calls):
        trial_number = trial + 1

        print(f"\n{'─' * 65}")
        print(f"  TRIAL {trial_number} of {opt.n_calls}")
        print(f"{'─' * 65}")

        # Step 1: Bayesian suggests widths → write directly to colormixcsv.xlsx
        phase = "PRESET (5,5,5)" if trial == 0 else "GP-GUIDED"
        print(f"\n  [Step 1 | Bayesian | {phase}] Selecting reservoir widths...")
        w1, w2, w3 = opt.ask()

        # Write to the xlsx that csvvolcont reads — keeps optimizer and chip in sync
        write_widths_to_xlsx(w1, w2, w3)

        total = w1 + w2 + w3
        pct = lambda w: f"{w/total*100:.1f}%" if total > 0 else "—"
        print(f"\n  ┌─ Suggested widths ──────────────────────┐")
        print(f"  │  piece_1 width : {w1:>3}  ({pct(w1)} of mix)    │")
        print(f"  │  piece_2 width : {w2:>3}  ({pct(w2)} of mix)    │")
        print(f"  │  piece_3 width : {w3:>3}  ({pct(w3)} of mix)    │")
        print(f"  │  total         : {total:>3}                     │")
        print(f"  └─────────────────────────────────────────┘")

        # Step 2: Mix on chip — reads widths from xlsx, connection already open
        print(f"\n  [Step 2] Running csvvolcont (mixing drop on chip)...")
        csvvolcont.main()

        # Step 3: Camera captures color
        print(f"\n  [Step 3] Capturing mixed color...")
        measured_hex = capture_color(cam)

        # Step 4: Bayesian scores the result
        print(f"\n  [Step 4] Scoring measured color against target...")
        converged = opt.tell(measured_hex)

        result_so_far = opt.get_result()
        this_delta_e  = opt._history[-1].delta_e

        print(f"\n  ┌─ Scoring ───────────────────────────────┐")
        print(f"  │  Target hex    : {target.hex:<24}  │")
        print(f"  │  Measured hex  : {measured_hex:<24}  │")
        print(f"  │  DeltaE        : {this_delta_e:<6.2f}  (this trial)      │")
        print(f"  │  Best DeltaE   : {result_so_far.best_delta_e:<6.2f}  (best so far)    │")
        print(f"  │  Converged     : {str(converged):<24}  │")
        print(f"  └─────────────────────────────────────────┘")
        print(f"\n  How DeltaE works:")
        print(f"    0–2   = visually identical  ✓")
        print(f"    2–10  = noticeable difference")
        print(f"    10+   = very different colors")

        # Stop early if converged — no cleanreload needed
        if converged:
            print(f"\n  [CONVERGED] DeltaE < 2.0 — match found after {trial_number} trial(s)!")
            break

        # Step 5: Move drop to graveyard, reload reservoirs
        print(f"\n  [Step 5] Moving drop to graveyard and reloading reservoirs...")
        cleanreload.move_to_graveyard(trial_number=trial_number)
        cleanreload.reload_reservoirs()

        if trial < opt.n_calls - 1:
            print(f"\n  Bayesian will use this score to suggest better widths next trial.")

    # ── Final result ────────────────────────────────────────────────────────
    result = opt.get_result()

    print(f"\n{'=' * 65}")
    print(f"  EXPERIMENT COMPLETE")
    print(f"{'=' * 65}")
    print(f"  Best widths found:")
    print(f"    piece_1 = {result.width_1}")
    print(f"    piece_2 = {result.width_2}")
    print(f"    piece_3 = {result.width_3}")
    print(f"  Best DeltaE  : {result.best_delta_e:.2f}")
    print(f"  Converged    : {result.converged}")
    print(f"  Target hex   : {result.target_hex}")
    print(f"\n  Full trial history saved to: optimization_log.csv")
    print(f"{'=' * 65}\n")

    opt.save_log(result)


if __name__ == "__main__":
    main()